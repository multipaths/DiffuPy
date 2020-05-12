# -*- coding: utf-8 -*-

"""Miscellaneous utils of the package."""
import itertools
import json
import logging
import pickle
import random
import warnings
from typing import List, Union, Dict, Optional

import networkx as nx
import numpy as np
import openpyxl as opxl
import pandas as pd
import pybel
from networkx import Graph

from .constants import *
from .constants import CSV, TSV, GRAPH_FORMATS

log = logging.getLogger(__name__)

"""Matrix/graph handling utils."""


def get_laplacian(graph: Graph, normalized: bool = False) -> np.ndarray:
    """Return Laplacian matrix."""
    if nx.is_directed(graph):
        warnings.warn('Since graph is directed, it will be converted to an undirected graph.')
        graph = graph.to_undirected()

    # Normalize matrix
    if normalized:
        return nx.normalized_laplacian_matrix(graph).toarray()

    return nx.laplacian_matrix(graph).toarray()


def set_diagonal_matrix(matrix: np.ndarray, d: list) -> np.ndarray:
    """Set diagonal matrix."""
    for j, row in enumerate(matrix):
        for i, x in enumerate(row):
            if i == j:
                matrix[j][i] = d[i]
            else:
                matrix[j][i] = x
    return matrix


def get_label_node(node: pybel.dsl.BaseAbundance) -> str:
    """Get label node."""
    if hasattr(node, 'name') and node.name is not None:
        if node.name.lower() == "":
            log.debug(f'Empty attribute name: {node.as_bel()}')
            return node.as_bel()

        else:
            return node.name.lower()

    elif hasattr(node, 'id') and node.id is not None:
        if node.id.lower() == "":
            log.debug(f'Empty attribute id: {node.as_bel()}')
            return node.as_bel()

        else:
            log.debug('Node labeled with id.' + node.id.lower())
            return node.id.lower()

    else:
        if node.as_bel() == "":
            log.debug('Node with no info.')
        else:
            log.debug(f'Node name nor id not labeled: {node.as_bel()}')
            return node.as_bel()


def get_label_list_graph(graph: nx.Graph, label: str) -> List:
    """Return graph labels."""
    if isinstance(graph, pybel.BELGraph):
        labels = []
        for node, _ in graph.nodes(data=True):
            labels.append(get_label_node(node))

        return labels

    elif nx.get_node_attributes(graph, label).values():
        return [
            value
            for value in nx.get_node_attributes(graph, label).values()
        ]

    elif graph.nodes:
        return graph.nodes()

    raise Warning('Could not get a label list from graph.')


def get_repeated_labels(labels):
    """Get duplicate labels."""
    seen = set()
    rep = []
    for x in labels:
        if x in seen:
            rep.append(x)
        seen.add(x)

    return rep


def get_label_ix_mapping(labels):
    """Get label to mat index mappings."""
    return {label: i for i, label in enumerate(labels)}


def get_label_scores_mapping(labels, scores):
    """Get label to scores mapping."""
    return {label: scores[i] for i, label in enumerate(labels)}


def get_idx_scores_mapping(scores):
    """Get mat index to scores mapping."""
    return {i: score for i, score in enumerate(scores)}


def map_intersection_type_background(background_labels: Dict[str, list], input_labels: list):
    """Intersection mapping."""
    labels_dict = {}

    for bck_label, bck_entities in background_labels.items():
        labels_dict[bck_label] = set(background_labels[bck_label]).intersection(input_labels)

    return labels_dict


def print_dict_dimensions(entities_db, title='Title', message=''):
    """Print dimension of the dictionary."""
    total = 0
    m = f'{title}\n'
    for k1, v1 in entities_db.items():
        m += f'\n{message}{k1}:\n'
        if isinstance(v1, dict):
            for k2, v2 in v1.items():
                m += f'{k2}  ({v2})\n'
        else:
            m += f'{v1}'

    print(f'{m}\n\n')


def log_dict(dict_to_print: dict, message: str = ''):
    """Print dictionary as list with a message."""
    for k1, v1 in dict_to_print.items():
        log.info(f'{message} {k1}: {v1} ')
        print(f'{message} {k1}: {v1} ')


def get_random_key_from_dict(d: dict) -> [Union[str, int, tuple]]:
    """Return random key from provided dict."""
    return random.choice(list(d.keys()))


def get_random_value_from_dict(d: dict):
    """Return random value from provided dict."""
    return d[get_random_key_from_dict(d)]


def lists_combinations(list_1, list_2):
    """Return all string combination from two list of strings."""
    return [x[0] + ' ' + x[1] for x in itertools.product(list_1, list_2)]


"""File loading/writting utils."""


def format_checker(fmt: str, fmt_list: list = GRAPH_FORMATS) -> None:
    """Check formats."""
    if fmt not in fmt_list:
        raise ValueError(
            f'The selected sep {fmt} is not valid. Please ensure you use one of the following formats: '
            f'{fmt_list}'
        )


def from_dataframe_file(path: str, fmt: str) -> pd.DataFrame:
    """Read network file."""
    format_checker(fmt)

    return pd.read_csv(
        path,
        header=0,
        sep=FORMAT_SEPARATOR_MAPPING[CSV] if fmt == CSV else FORMAT_SEPARATOR_MAPPING[TSV]
    )


def from_json(path: str):
    """Read from json file."""
    with open(path) as f:
        return json.load(f)


def to_json(data, path: str):
    """Save json file."""
    with open(path, 'w') as f:
        json.dump(data, f, indent=2)


def from_pickle(input_path):
    """Read from pickle file."""
    with open(input_path, 'rb') as f:
        unpickler = pickle.Unpickler(f)
        return unpickler.load()


def from_nparray_to_df(nparray: np.ndarray) -> pd.DataFrame:
    """Convert numpy array to data frame."""
    return pd.DataFrame(data=nparray[1:, 1:],
                        index=nparray[1:, 0],
                        columns=nparray[0, 1:])


"""Data parsing utils."""


def decode_labels(labels):
    """Validate labels."""
    labels_decode = []

    for label in labels:
        if not isinstance(label, str):

            if isinstance(label, int):
                label = str(label)
            else:
                label = label.decode('utf-8').replace('"', '')

        labels_decode.append(label)

    return labels_decode


def munge_label(label: Union[str, int, float]) -> str:
    """Munge label strings."""
    remove_set = ['*', ' ', '|', '-', '"', "'", "↑", "↓", "\n"]
    split_set = ['/']

    label = str(label).lower()

    for symb in remove_set:
        if symb in label:
            label = label.replace(symb, '')

    for symb in split_set:
        if symb in label:
            label = tuple(set(label.split(symb)))
            if len(label) == 1:
                label = label[0]

    return label


def munge_label_list(labels: list):
    """Munge labels list."""
    return list(set([munge_label(label) for label in labels]))


def munge_label_scores_dict(labels: dict) -> Dict[str, Union[list, int, str]]:
    """Munge labels dict."""
    return {munge_label(label): v for label, v in labels.items()}


def munge_label_type_dict(label_dict: Dict[str, Union[list, int, str, dict]]) -> Dict[str, Union[list, int, str, dict]]:
    """Munge labels type dict."""
    type_label_dict = {}

    for type_label, labels in label_dict.items():
        if isinstance(labels, dict):
            type_label_dict[type_label] = munge_label_scores_dict(labels)

        elif isinstance(labels, dict):
            type_label_dict[type_label] = munge_label_scores_dict(labels)

    return type_label_dict


def munge_cell(cell):
    """Munge cell."""
    if isinstance(cell, str):
        if cell.replace(',', '').replace('.', '').replace('-', '').isnumeric():
            return float(cell)
        else:
            return munge_label(cell)

    elif isinstance(cell, float) or isinstance(cell, int):
        return cell

    elif cell is None:
        return 'NA'

    else:
        raise TypeError(f'The cell "{cell}" could not be processed.')


def parse_xls_sheet_to_df(sheet: opxl.workbook,
                          min_row: Optional[int] = 1,
                          relevant_cols: Optional[list] = None,
                          irrelevant_cols: Optional[list] = None) -> pd.DataFrame:
    """Process/format excel sheets to DataFrame."""
    parsed_sheet_dict = {}

    for col in sheet.iter_cols(min_row=min_row):
        col_label = col[0].value

        if ((relevant_cols is not None and col_label in relevant_cols) or (
                irrelevant_cols is not None and col_label not in irrelevant_cols)):
            parsed_sheet_dict[col_label] = [munge_cell(cell.value) for cell in col[1:]]

    return pd.DataFrame.from_dict(parsed_sheet_dict)


def parse_xls_to_df(path: str,
                    min_row: Optional[int] = 1,
                    relevant_sheets: Optional[list] = None,
                    irrelevant_sheets: Optional[list] = None,
                    relevant_cols: Optional[list] = None,
                    irrelevant_cols: Optional[list] = None,
                    ) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Process excel file as a set (if several excel sheets) or a single dataframe."""
    wb = opxl.load_workbook(filename=path)

    sheets = wb.sheetnames

    if len(sheets) > 1:
        return {sheets[ix].lower(): parse_xls_sheet_to_df(sheet, min_row, relevant_cols, irrelevant_cols)
                for ix, sheet in enumerate(wb)
                if (relevant_sheets is not None and sheets[ix] in relevant_sheets) or (
                    irrelevant_sheets is not None and sheets[ix] not in irrelevant_sheets)
                }

    else:
        return parse_xls_sheet_to_df(wb[sheets[0]])
